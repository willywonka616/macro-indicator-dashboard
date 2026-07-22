"""Gold-inclusive reserves (no API key on either source).

Resolves STATUS.md §5's flagged reserves discrepancy: `TRESEGUSM052N` is
"Total Reserves EXCLUDING Gold" — confirmed by its own name and FRED
metadata — so it structurally cannot reconcile with Dalio's Ch.17 "FX
reserves, 3% of GDP" figure. US reserves excl. gold are ~0.8% of GDP; only
adding gold **at market price** (not the $42.2222/oz statutory book value
Treasury carries it at) closes the gap, because the US holds a large gold
stock (~261.5M fine troy ounces) that is worth roughly $800B+ at market vs.
~$11B at the statutory rate.

Two live, no-key sources, composed here (checked in this order of
preference before implementing):
  1. An off-the-shelf "reserves including gold" series (World Bank
     FI.RES.TOTL.CD, sourced from IMF IFS) exists, but World Bank WDI
     indicators lag 1-2 years — too stale for a dashboard that refreshes
     monthly. IMF IFS itself likely has a more current series
     (something like RAFA_USD) but the exact US country-code key could not
     be confirmed without live API access. Neither is used here; this is
     recorded so a future session can pick the investigation back up instead
     of re-deriving it composed if a good current source turns up.
  2. What's actually used: Treasury's own gold holdings (fine troy ounces,
     monthly) x a live gold price, combined with FRED's excl.-gold reserves
     and GDP. Both endpoints are exercised by verify() so a schema change is
     visible in the run log, same as every other integration in this
     project.

Gold-price source history (2026-07): DBnomics' LBMA/gold_D dataset — the
obvious first choice, and what this module used originally — now 404s on
every shape tried (series-code path, dataset-level with a `dimensions`
filter, and a bare unfiltered dump). Root cause, confirmed via web research
rather than guessed: LBMA moved its benchmark price tables behind a
members-only portal starting the week of 2025-11-24, and FRED's own
GOLDAMGBD228NLBM series (the same LBMA fix) was separately discontinued
with no replacement — DBnomics' LBMA mirror is downstream of the same
now-restricted source, not a URL-shape problem. Switched to IMF's Primary
Commodity Price System (PCPS, indicator code PGOLD — world gold price
benchmark, USD/troy oz, monthly), mirrored on DBnomics the same way COFER
is, which keeps the pipeline on http-only, no-key sources dumped defensively
(bare fetch + client-side filter by series_code, since PCPS's exact
dimension names haven't been confirmed live yet — see verify()).

Best-effort by design: gold_market_value_usd() raises on any problem, and the
caller (fetch.py) falls back to a manual value rather than shipping a
live-tagged number built on a stale price, same pattern as imf.py's COFER
fallback.

**Search for a fresher price (2026-07, TASKgoldpricefreshness.md): all five
candidates tried, none currently usable, order as specified:**
  1. FRED (`series/search?search_text=gold price`, plus `LBMA gold`/`gold
     fixing`/`gold spot`) — zero results for any actual spot/fixing series.
     `GOLDAMGBD228NLBM`/`GOLDPMGBD228NLBM` (the old LBMA fixings) now
     return a hard 400, not just "discontinued with history" — the IDs are
     gone outright. The only current, on-cadence hits are a volatility
     index (GVZCLS) and PPI/import/export price INDICES (base-year,
     e.g. `IQ12260`), none of them a $/oz spot price.
  2. World Bank Pink Sheet via DBnomics (provider `WB`) — the real dataset
     (`commodity_prices`) exists and has a gold series (`FGOLD-1W`), but
     it's annual with `2030` (a *projected* out-year) as its "latest"
     point, not the monthly Pink Sheet. See docs/review/2026-07-20c.
  3. Bundesbank, both directly (api.statistiken.bundesbank.de, three
     guessed series keys — all 404) and via DBnomics (provider `BUBA`):
     found the real gold series (`BBEX3`, D-Mark Frankfurt fixing) after
     a full-catalog + server-side search, but it stops at end-1998 —
     pre-euro, permanently discontinued, not a stale-but-alive feed.
  4. Nasdaq Data Link (ex-Quandl), LBMA/GOLD, no-key attempt — blocked by
     an Incapsula anti-bot WAF (403, JS challenge page), same failure mode
     Stooq hit below. Would need a paid API key (a new GitHub secret) to
     even test properly; not pursued without that being requested.
  5. Stooq daily XAUUSD CSV, no key — a bare request 404s; with a
     browser User-Agent it returns 200 but the body is a client-side
     SHA-256 proof-of-work anti-bot challenge, not the CSV. Not reachable
     via a plain HTTP client.

No source switch made — DBnomics-mirrored IMF PCPS remains primary, manual
value remains the fallback, unchanged. What DID change: fetch.py now runs a
freshness check on the fetched price (see series.py's `require_fresh`)
before trusting it as live, so a frozen PCPS feed like the current one
(stuck at 2025-06) correctly degrades to the manual value instead of
shipping as falsely "live" — closing the actual gap this task opened with,
even without a fresher source to switch to.

**Automated for real (2026-07, TASKgoldautomation.md) — DBnomics/IMF PCPS
removed, replaced with the World Bank's own "Pink Sheet" data, live:**

  §1 retried Nasdaq Data Link and Stooq with a full browser header set
  (User-Agent, Accept, Accept-Language) rather than default-`requests`
  headers, per the task's hypothesis that the earlier failures were
  fingerprinting, not IP-based. Both still fail identically to before —
  Nasdaq returns Incapsula's 403 challenge page, Stooq returns its
  client-side proof-of-work challenge — with headers changing nothing.
  Genuinely blocked, not a header artifact. See verify()'s retried probes
  for the exact response each run.

  §2 LBMA's own feed (`prices.lbma.org.uk/json/gold_*.json`) 403s outright —
  confirms the module docstring's earlier finding: LBMA moved historical/
  tabulated data behind its licensed MyLBMA portal; the public JSON feed
  path that used to work is gone, not just DBnomics' mirror of it.

  §3 World Bank Pink Sheet, direct download (not DBnomics): the actual
  monthly "CMO Historical Data" spreadsheet, not the annual
  Commodity Markets Outlook *forecast* table DBnomics exposed. Implemented
  as `_worldbank_pink_sheet_gold_monthly()` — downloaded straight from
  thedocs.worldbank.org with browser headers and parsed defensively
  (searches for a "gold" column header and a `YYYY`M`MM`-labelled date
  column rather than a hardcoded cell range, since the exact layout could
  not be confirmed live: thedocs.worldbank.org is blocked by this
  project's own dev-sandbox proxy, AND returned a bare 403 to an
  independent out-of-sandbox fetch attempt during research — sits behind
  bot/CDN protection). Kept as the primary path per the task's explicit
  preference for the direct download; verify() dumps exactly what it finds
  (or the exact failure) on every run so this is confirmed or corrected
  live rather than assumed.

  §4's "maintained GitHub dataset mirror" suggestion turned out to be the
  practical winner: `datasets/gold-prices` — a GitHub repo that mirrors
  this *exact same* World Bank Commodity Markets data, auto-updated daily
  via its own GitHub Actions workflow (confirmed current live during
  research: latest observation 2026-06, and its own README/datapackage.json
  document the World Bank as the 1960-present source). Reachable even from
  this project's restrictive dev sandbox — same host
  (raw.githubusercontent.com) cbo.py already relies on — so it doubles as
  a live-tested fallback leg when the direct download is blocked. Not a
  different or lesser source, just a different transport for the same
  institutional data; used as `_github_mirror_gold_monthly()`, second in
  `gold_price_usd_per_oz()`'s try order.

  Other §4 candidates, probed and not used: SNB's data portal
  (`data.snb.ch`) publishes Switzerland's own gold bullion *holdings*
  (analogous to this project's existing Treasury gold-ounces series), not
  a market price benchmark — a scope mismatch, not a blocked source.
  UNCTADstat's public site 403s to an out-of-sandbox fetch and its
  documented access pattern is a session-based SDMX portal, not a
  lightweight bulk endpoint; not pursued once §3/§4's Pink Sheet path
  worked. FX-style "XAU as a currency" APIs (exchangerate.host, Metals-API,
  GoldAPI.io, Commodities-API, UniRateAPI) turned out to all require a key
  in their free tier as of 2026 — none is actually keyless despite being
  commonly described that way.

  §5 (accepting a keyed source) not needed — §3/§4 produced a working
  keyless institutional source.

  Net effect: `gold_price_usd_per_oz()` no longer touches DBnomics/IMF PCPS
  at all (that code is removed, not just unused — it was permanently
  frozen since 2025-06 per the freshness guard, see STATUS.md §16/§17).
  `manual_price` should no longer fire for gold under normal operation;
  `goldPriceManualFallback` in data/manual.json stays wired in fetch.py as
  the last-resort leg if both World Bank paths ever go down together.

**Follow-up (2026-07-22, same day): a monthly average is not spot, even
when its date label passes a freshness check.** The round above shipped
correctly, but a manual price check on the shipped figure found it
implausible: reserves landed at 4.9% of GDP off a $4,228/oz "June 2026"
Pink Sheet average, while actual spot in the days around when this was
checked was running $4,070-4,083/oz (~3.5% lower) — a real, material gap,
not a rounding artifact. Root cause is structural, not a bug in the
freshness guard: the Pink Sheet's monthly figure is an AVERAGE across the
whole month (confirmed by cross-referencing daily spot prices for June
2026 — the month opened near $4,460, fell through the month to ~$4,005 by
June 30, and a $4,228 whole-month average is consistent with that
trajectory), so even a "fresh" (recently-dated) observation represents
gold's price as of roughly the middle of the averaging window, not today —
and gold moves enough in a few weeks for that gap to matter, unlike the
other monthly/quarterly series in this pipeline. Added LBMA's own daily
PM-fix feed (`_lbma_gold_monthly_and_latest()`) as the new primary leg,
ahead of the World Bank Pink Sheet: it's the actual published gold
benchmark (not a redistributor), updates on every London business day,
and this module now keeps the LATEST DAY seen in each month (not an
average) specifically so the current month's bucket is as close to spot
as the feed allows. The World Bank Pink Sheet (direct, then GitHub
mirror) is now the fallback tier, unchanged otherwise. See STATUS.md §23.
"""

from __future__ import annotations

import datetime as dt
import io
import re
import time
from collections import defaultdict

import openpyxl
import requests

import series as S

TREASURY_BASE = "https://api.fiscaldata.treasury.gov/services/api/fiscal_service"
GOLD_ENDPOINT = "/v2/accounting/od/gold_reserve"

# World Bank Commodity Markets ("Pink Sheet"), direct download — the actual
# monthly historical spreadsheet, not the annual forecast table DBnomics
# exposed (see module docstring). URL carries a per-publication hash
# component, same pattern as CBO's recurring-data URLs.
#
# Confirmed live (2026-07) that this rotates WITHOUT the old URL ever
# breaking: a first hash (".../18675f1d...-0050012025/...") kept resolving
# 200 and parsing cleanly, but its underlying data had frozen at 2025-12 —
# seven months stale — while the World Bank had already moved on to a new
# hash (".../74e8be41...-0050012026/...", in place since 2026-02-03,
# confirmed current). A successful parse is therefore NOT proof of
# freshness here, unlike a 404 would be — see gold_price_usd_per_oz_labeled(),
# which checks the direct leg's own latest date and prefers the GitHub
# mirror whenever it's stale, so a future hash rotation degrades
# automatically instead of silently shipping old data or requiring another
# manual URL fix.
WORLDBANK_PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/74e8be41ceb20fa0da750cda2f6b9e4e-"
    "0050012026/related/CMO-Historical-Data-Monthly.xlsx"
)

# Nasdaq/Stooq/LBMA all reject a bare python-requests User-Agent outright
# (WAF/bot-challenge — see module docstring's retried probes); a
# browser-shaped header set is cheap insurance against the same
# fingerprinting hitting thedocs.worldbank.org, which also sits behind a
# CDN and returned a bare 403 to at least one out-of-sandbox fetch attempt
# during research.
_BROWSER_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    "Accept": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
               "text/csv,application/json,*/*;q=0.8"),
    "Accept-Language": "en-US,en;q=0.9",
}

# GitHub mirror of the same World Bank Pink Sheet data (TASKgoldautomation.md
# §4's "maintained GitHub dataset mirror") — auto-updated daily per its own
# README, confirmed live and current during research (latest observation
# 2026-06). Reachable even from this project's dev sandbox, same host
# (raw.githubusercontent.com) cbo.py already relies on. Used as the fallback
# leg, not the primary — see gold_price_usd_per_oz().
GOLD_MIRROR_REPO = "datasets/gold-prices"
GOLD_MIRROR_RAW_URL = f"https://raw.githubusercontent.com/{GOLD_MIRROR_REPO}/main/data/monthly.csv"

# Row labels in the Pink Sheet's date column look like "1960M01".
_YM_CELL_RE = re.compile(r"^\s*(\d{4})M(\d{2})\s*$")

# LBMA's own public price feed — the actual published gold benchmark (not
# a redistributor), daily, so its most recent entry is normally only a
# business day or two old. Preferred over the World Bank Pink Sheet
# (§22.3/§22.4) for THIS row specifically: a follow-up round
# (2026-07-22b) found that a monthly AVERAGE — even one that passes a
# freshness check on its date label — can sit weeks behind actual spot for
# a metal this volatile (confirmed live: the shipped June 2026 Pink Sheet
# average, $4,228/oz, was ~3.5% above the ~$4,070-4,083/oz spot range
# observed in late July 2026, moving in the direction of gold's earlier
# 2026 highs, not a small rounding gap). LBMA's daily fix doesn't have
# that structural lag. See module docstring's earlier round for why this
# feed was initially assumed still fully behind a licensed portal — a
# retry with browser headers found the public JSON endpoint still
# resolves; this round implements it for real rather than leaving it a
# diagnostic-only probe.
LBMA_GOLD_PM_URL = "https://prices.lbma.org.uk/json/gold_pm.json"
LBMA_FRESH_DAYS = S.FRESHNESS_DAYS_BY_FREQ["Daily"]


# --- http ------------------------------------------------------------------

def _get_json(url: str, params: dict, tries: int = 4):
    last = None
    for i in range(tries):
        try:
            r = requests.get(url, params=params, timeout=45)
            r.raise_for_status()
            return r.json()
        except requests.exceptions.RequestException as e:  # noqa: PERF203
            last = e
            if i < tries - 1:
                time.sleep(2 ** i)  # 1s, 2s, 4s
    raise RuntimeError(f"request to {url} failed after {tries} tries: {last}")


def _pick(sample: dict, names, contains=None, exclude=()):
    for n in names:
        if n in sample:
            return n
    if contains:
        for k in sample:
            kl = k.lower()
            if all(t in kl for t in contains) and not any(e in kl for e in exclude):
                return k
    return None


def _num(x):
    try:
        return float(str(x).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _ym(record_date: str):
    d = dt.date.fromisoformat(record_date)
    return (d.year, d.month)


# --- Treasury: gold holdings, fine troy ounces ------------------------------

def _treasury_gold_rows():
    rows, page = [], 1
    out = []
    while page <= 20:
        js = _get_json(f"{TREASURY_BASE}{GOLD_ENDPOINT}",
                        {"format": "json", "page[size]": "1000", "page[number]": str(page),
                         "sort": "record_date"})
        data = js.get("data", [])
        out.extend(data)
        meta = js.get("meta", {})
        total_pages = int(meta.get("total-pages") or meta.get("total_pages") or 1)
        if page >= total_pages or not data:
            break
        page += 1
    if not out:
        raise RuntimeError("gold_reserve: no rows returned")
    return out


def gold_holdings_troy_oz() -> dict:
    """{(year, month): total fine troy ounces held} — monthly.

    Sums across every row for a given date (locations/categories) rather than
    assuming a single pre-summed "total" row exists — safer if the schema has
    a location breakdown with no grand-total row, and harmless (same result)
    if it does have one under a different name we didn't guess.
    """
    rows = _treasury_gold_rows()
    s = rows[0]
    ozk = _pick(s, ["fine_troy_ounce_qty", "fine_troy_ounce_amt"],
                contains=["troy", "oz"]) or _pick(s, [], contains=["troy"])
    if not ozk:
        raise RuntimeError(f"gold_reserve: no troy-ounce field in {list(s)}")
    totk = _pick(s, ["location_desc", "summary_fund_type_desc"], contains=["desc"])

    monthly = defaultdict(float)
    for row in rows:
        # if a grand-total row is identifiable, prefer it exclusively per
        # month so we don't double count; otherwise sum every row.
        if totk and "total" in str(row.get(totk, "")).lower():
            oz = _num(row.get(ozk))
            if oz is not None:
                monthly[_ym(row["record_date"])] = oz  # overwrite: total wins
    if monthly:
        return dict(monthly)

    monthly = defaultdict(float)
    for row in rows:
        oz = _num(row.get(ozk))
        if oz is not None:
            monthly[_ym(row["record_date"])] += oz
    return {k: v for k, v in monthly.items() if v}


# --- LBMA: the actual gold benchmark, daily, closest to spot ---------------

def _lbma_gold_monthly_and_latest():
    """({(year, month): USD/oz}, latest observation date) from LBMA's own
    public PM-fix feed. Buckets by month (to match the rest of this
    module's {(y, m): price} interface) using the LATEST day observed in
    each month — not a monthly average — so the current month's bucket is
    as close to spot as the feed allows. Returns the actual latest `date`
    too (not just its month) because a same-day observation still counts
    as "this month" days after it happened; using month-start to gauge
    freshness would either overstate staleness for a genuinely fresh feed
    or (if using month-end) understate it for a genuinely dead one — the
    caller needs the real day.
    """
    r = requests.get(LBMA_GOLD_PM_URL, headers=_BROWSER_HEADERS, timeout=45)
    r.raise_for_status()
    rows = r.json()
    if not isinstance(rows, list) or not rows:
        raise RuntimeError(f"{LBMA_GOLD_PM_URL}: unexpected/empty response shape")

    monthly: dict = {}  # (y, m) -> (day, usd) — keep the LATEST day per month
    latest_date = None
    for row in rows:
        d, v = row.get("d"), row.get("v")
        if not d or not isinstance(v, list) or not v or v[0] is None:
            continue
        try:
            y, m, day = (int(x) for x in d.split("-"))
            usd = float(v[0])  # confirmed live (2026-07): v[0] is USD/oz —
        except (TypeError, ValueError):              # e.g. 37.7 for 1968-04-01,
            continue                                  # matching the known historical fixing
        key = (y, m)
        if key not in monthly or day > monthly[key][0]:
            monthly[key] = (day, usd)
        d_obj = dt.date(y, m, day)
        if latest_date is None or d_obj > latest_date:
            latest_date = d_obj
    if not monthly or latest_date is None:
        raise RuntimeError(f"{LBMA_GOLD_PM_URL}: no usable USD observations")
    return {k: val[1] for k, val in monthly.items()}, latest_date


# --- World Bank Pink Sheet: direct download ---------------------------------

def _get_bytes(url: str, headers: dict, tries: int = 3):
    last = None
    for i in range(tries):
        try:
            r = requests.get(url, headers=headers, timeout=45)
            r.raise_for_status()
            return r.content
        except requests.exceptions.RequestException as e:  # noqa: PERF203
            last = e
            if i < tries - 1:
                time.sleep(2 ** i)
    raise RuntimeError(f"request to {url} failed after {tries} tries: {last}")


def _worldbank_pink_sheet_gold_monthly() -> dict:
    """{(year, month): USD per troy oz}, parsed from the World Bank's own
    Pink Sheet spreadsheet (direct download — TASKgoldautomation.md §3's
    explicit preference over a mirror). The exact sheet/column layout is
    unconfirmed from this project's dev sandbox (see module docstring), so
    this is parsed defensively rather than against a hardcoded cell range:
    find whichever sheet looks like the monthly-prices one (falling back to
    the first sheet), scan its first few rows for a header cell containing
    "gold" but not "index", then read down column A for date labels shaped
    like "1960M01" next to it. Raises with a specific, actionable message
    at each stage so a real layout change is easy to diagnose from the run
    log rather than silently returning nothing.
    """
    content = _get_bytes(WORLDBANK_PINK_SHEET_URL, _BROWSER_HEADERS)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    sheet = None
    for name in wb.sheetnames:
        low = name.lower()
        if "monthly" in low and "price" in low:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    gold_col = None
    header_row_idx = None
    for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=8), start=1):
        for cell in row:
            v = str(cell.value or "").strip().lower()
            if "gold" in v and "index" not in v:
                gold_col = cell.column
                header_row_idx = r_idx
                break
        if gold_col:
            break
    if gold_col is None:
        raise RuntimeError(
            f"worldbank pink sheet: no 'gold' column header found in sheet "
            f"{sheet.title!r} (sheets present: {wb.sheetnames})")

    monthly = {}
    for row in sheet.iter_rows(min_row=header_row_idx + 1):
        label = row[0].value if row else None
        m = _YM_CELL_RE.match(str(label)) if label else None
        if not m or gold_col - 1 >= len(row):
            continue
        val = row[gold_col - 1].value
        if val is None:
            continue
        try:
            monthly[(int(m.group(1)), int(m.group(2)))] = float(val)
        except (TypeError, ValueError):
            continue
    if not monthly:
        raise RuntimeError(
            f"worldbank pink sheet: found a 'gold' header at column {gold_col} "
            f"(row {header_row_idx}) in sheet {sheet.title!r} but no dated rows under it")
    return monthly


# --- GitHub mirror of the same data: fallback leg ---------------------------

def _github_mirror_gold_monthly() -> dict:
    """{(year, month): USD per troy oz} from `datasets/gold-prices` — see
    module docstring for why this is trusted (same underlying World Bank
    data, confirmed live/current, auto-updated daily)."""
    r = requests.get(GOLD_MIRROR_RAW_URL, timeout=30)
    r.raise_for_status()
    monthly = {}
    lines = r.text.strip().splitlines()
    for line in lines[1:]:  # skip "Date,Price" header
        parts = line.split(",")
        if len(parts) != 2:
            continue
        ym, val = parts
        try:
            y, m = ym.split("-")
            monthly[(int(y), int(m))] = float(val)
        except (ValueError, TypeError):
            continue
    if not monthly:
        raise RuntimeError(f"{GOLD_MIRROR_RAW_URL}: no usable rows parsed")
    return monthly


def gold_price_usd_per_oz_labeled() -> tuple[dict, str]:
    """Same as gold_price_usd_per_oz(), but also returns which leg actually
    served the data — fetch.py uses the label so a row's `src` names the
    real leg rather than a generic name that would hide which one is live
    this run.

    Three legs, in order:
      1. LBMA's own daily PM fix — closest to spot (see LBMA_GOLD_PM_URL's
         comment: a monthly average, even a "fresh" one by date label, can
         sit weeks behind actual spot for a metal this volatile).
      2. World Bank Pink Sheet, direct download — only if its OWN latest
         observation is itself fresh, since a successful parse is not
         proof of freshness for this source either (the direct download
         can keep resolving 200 for months after the World Bank quietly
         rotates its live data to a new URL hash, confirmed live 2026-07).
      3. The GitHub mirror of the same World Bank data.
    Each leg is skipped (not just on exception, but also when it parses
    fine but is itself stale) so a future outage or silent staleness on
    any leg degrades automatically to the next, rather than shipping old
    data or needing a manual fix.
    """
    try:
        lbma, lbma_latest = _lbma_gold_monthly_and_latest()
        age_days = (dt.date.today() - lbma_latest).days
        if age_days <= LBMA_FRESH_DAYS:
            return lbma, "LBMA (PM fix, direct)"
        print(f"LBMA gold_pm.json parsed fine but its latest observation "
              f"({lbma_latest.isoformat()}) is {age_days}d old — trying "
              f"the World Bank Pink Sheet instead")
    except Exception as e:  # noqa: BLE001
        print(f"LBMA direct feed failed, trying the World Bank Pink Sheet: {e}")

    direct, direct_err = None, None
    try:
        direct = _worldbank_pink_sheet_gold_monthly()
    except Exception as e:  # noqa: BLE001
        direct_err = e

    if direct:
        latest = max(direct)
        age_days = (dt.date.today() - dt.date(latest[0], latest[1], 1)).days
        if age_days <= S.FRESHNESS_DAYS_BY_FREQ["Monthly"]:
            return direct, "World Bank (Pink Sheet, direct)"
        print(f"World Bank Pink Sheet direct download parsed fine but its latest "
              f"observation ({latest[0]}-{latest[1]:02d}) is {age_days}d old — likely "
              f"a stale per-publication URL hash (see module docstring); trying "
              f"the GitHub mirror instead")
    else:
        print(f"World Bank Pink Sheet direct download failed, trying GitHub mirror: {direct_err}")

    return (_github_mirror_gold_monthly(),
            f"World Bank (Pink Sheet, GitHub mirror: {GOLD_MIRROR_REPO})")


def gold_price_usd_per_oz() -> dict:
    """{(year, month): USD per troy oz}. Tries LBMA's daily feed first, then
    the World Bank Pink Sheet direct download, then its GitHub mirror — see
    gold_price_usd_per_oz_labeled() for the full three-leg order and the
    source-labelled version. If all three fail, raises — the caller
    (fetch.py) then falls further to the manual price input."""
    data, _label = gold_price_usd_per_oz_labeled()
    return data


# --- diagnostic-only probes (TASKgoldautomation.md §§1-2) -------------------
# Never used as a live source — Nasdaq/Stooq/LBMA are all documented dead
# ends (module docstring). Kept as best-effort, non-raising probes purely so
# every verify() run reports their exact current failure mode, per the
# task's "report exactly what you find" instruction, instead of relying on
# a stale prose claim that could silently go out of date.

def _probe(label: str, url: str, headers: dict | None = None, find: str | None = None):
    """`find`, if given, is searched for anywhere in the full body (not
    just the 200-char preview) and reported explicitly — a preview alone
    can miss a key buried in a large JSON object (e.g. "is XAU present in
    this currency list" can't be answered by the first 200 characters)."""
    try:
        r = requests.get(url, headers=headers or {}, timeout=20)
        preview = r.text[:200].replace("\n", " ")
        print(f"  [{label}] {url} -> HTTP {r.status_code}, "
              f"content-type={r.headers.get('content-type')}, body preview: {preview!r}")
        if find is not None:
            present = find in r.text
            print(f"  [{label}] {find!r} present anywhere in full body: {present}")
    except requests.exceptions.RequestException as e:
        print(f"  [{label}] {url} -> FAILED: {e}")


# --- combined ----------------------------------------------------------------

def gold_market_value_usd() -> dict:
    """{(year, month): USD market value of US gold holdings}."""
    oz = gold_holdings_troy_oz()
    price = gold_price_usd_per_oz()
    months = oz.keys() & price.keys()
    if not months:
        raise RuntimeError("gold: no overlapping months between holdings and price")
    return {k: oz[k] * price[k] for k in months}


# --- verification ------------------------------------------------------------

def verify() -> bool:
    """Dump every endpoint's schema + latest values, and try the computation.
    Non-fatal: always returns True (a source hiccup must not red the run —
    the build degrades through the fallback chain instead)."""
    print("\nVerifying gold sources (Treasury holdings + gold price: LBMA "
          "daily fix, then World Bank Pink Sheet direct, then its GitHub "
          "mirror; non-fatal — manual fallback on failure)\n")
    try:
        rows = _treasury_gold_rows()
        s = rows[0]
        print(f"[gold-holdings] {TREASURY_BASE}{GOLD_ENDPOINT}")
        print("  fields:", ", ".join(s))
        for col in [c for c in s if c.endswith("_desc")]:
            seen = []
            for r in rows:
                v = r.get(col)
                if v is not None and v not in seen:
                    seen.append(v)
                if len(seen) >= 15:
                    break
            print(f"  distinct {col}: {seen}")
        latest = max(r["record_date"] for r in rows)
        print(f"  latest {latest} rows:")
        for r in [r for r in rows if r["record_date"] == latest][:10]:
            print(f"    {r}")
        f = S.freshness("gold holdings (Treasury oz)", latest, S.FRESHNESS_DAYS_BY_FREQ["Monthly"])
        fresh_s = "STALE" if f["stale"] else "ok"
        print(f"  freshness: {f['age_days']}d old, {f['max_age_days']}d threshold, {fresh_s}")
    except Exception as e:  # noqa: BLE001
        print(f"[gold-holdings] FAILED: {e}")

    print("\n[gold-price] LBMA daily PM fix, direct — primary leg (closest to spot)")
    print(f"  {LBMA_GOLD_PM_URL}")
    try:
        lbma, lbma_latest = _lbma_gold_monthly_and_latest()
        print(f"  parsed {len(lbma)} months; latest observation: {lbma_latest.isoformat()} "
              f"= {lbma[max(lbma)]} USD/oz")
        age_days = (dt.date.today() - lbma_latest).days
        print(f"  freshness: {age_days}d old, {LBMA_FRESH_DAYS}d threshold, "
              f"{'STALE' if age_days > LBMA_FRESH_DAYS else 'ok'}")
    except Exception as e:  # noqa: BLE001
        print(f"  FAILED: {e}")

    print("\n[gold-price] World Bank Pink Sheet, direct download "
          "(TASKgoldautomation.md §3) — fallback leg")
    print(f"  {WORLDBANK_PINK_SHEET_URL}")
    try:
        direct = _worldbank_pink_sheet_gold_monthly()
        latest = max(direct)
        print(f"  parsed {len(direct)} months; latest: {latest[0]}-{latest[1]:02d} "
              f"= {direct[latest]} USD/oz")
        max_days = S.FRESHNESS_DAYS_BY_FREQ["Monthly"]
        f = S.freshness("gold price (World Bank direct)", f"{latest[0]}-{latest[1]:02d}-01", max_days)
        print(f"  freshness: {f['age_days']}d old, {max_days}d threshold, "
              f"{'STALE' if f['stale'] else 'ok'}")
    except Exception as e:  # noqa: BLE001
        print(f"  FAILED: {e}")

    print(f"\n[gold-price] GitHub mirror ({GOLD_MIRROR_REPO}), fallback leg")
    print(f"  {GOLD_MIRROR_RAW_URL}")
    try:
        mirror = _github_mirror_gold_monthly()
        latest = max(mirror)
        print(f"  parsed {len(mirror)} months; latest: {latest[0]}-{latest[1]:02d} "
              f"= {mirror[latest]} USD/oz")
        max_days = S.FRESHNESS_DAYS_BY_FREQ["Monthly"]
        f = S.freshness("gold price (GitHub mirror)", f"{latest[0]}-{latest[1]:02d}-01", max_days)
        print(f"  freshness: {f['age_days']}d old, {max_days}d threshold, "
              f"{'STALE' if f['stale'] else 'ok'}")
    except Exception as e:  # noqa: BLE001
        print(f"  FAILED: {e}")
    # Ask the real selector, not just "did the direct parse succeed" —
    # gold_price_usd_per_oz_labeled() also rejects a direct leg that parsed
    # fine but is itself stale (see its docstring), so this line reports
    # what actually ships, not just what didn't throw.
    try:
        _, active_label = gold_price_usd_per_oz_labeled()
        print(f"\n  active leg this run: {active_label}")
    except Exception as e:  # noqa: BLE001
        print(f"\n  active leg this run: NONE — all three legs failed: {e}")

    print("\n[retry, TASKgoldautomation.md §1] Nasdaq Data Link LBMA/GOLD, "
          "browser headers — diagnostic only, never used as a live source "
          "(see module docstring)")
    _probe("nasdaq", "https://data.nasdaq.com/api/v3/datasets/LBMA/GOLD.csv?rows=5", _BROWSER_HEADERS)

    print("\n[retry, §1] Stooq XAUUSD daily CSV, browser headers — diagnostic only")
    _probe("stooq", "https://stooq.com/q/d/l/?s=xauusd&i=d", _BROWSER_HEADERS)

    print("\n[probe, follow-up round] exchangerate-api.com open-access (keyless) "
          "tier — does XAU actually appear without a key? Marketing pages "
          "conflate the keyed 'Free' tier with the truly-open one; diagnostic "
          "only, never used as a live source unless this confirms XAU present")
    _probe("open.er-api.com", "https://open.er-api.com/v6/latest/USD", _BROWSER_HEADERS, find='"XAU"')

    try:
        mv = gold_market_value_usd()
        last = max(mv)
        print(f"\n  computed gold market value: ${mv[last]/1e9:.1f}B "
              f"as of {last[0]}-{last[1]:02d} ({len(mv)} months)")
    except Exception as e:  # noqa: BLE001
        print(f"\n  gold market value computation FAILED: {e}")
    return True
